import matplotlib.pyplot as plt
import arabic_reshaper
from bidi.algorithm import get_display
from openpyxl import load_workbook





# --- خواندن داده‌ها از اکسل ---
try:
    wb = load_workbook('data.xlsx')
    sheet = wb.active


    provinces = []
    simcards = []

    for row in sheet.iter_rows(min_row=2, values_only=True):  # از ردیف دوم شروع می‌کند
        provinces.append(row[0])  # ستون اول: استان‌ها
        simcards.append(row[1])  # ستون دوم: تعداد سیم‌کارت

except Exception as e:
    print(f"خطا در خواندن فایل اکسل: {e}")
    exit()

# --- محاسبه مجموع ---
total = sum(simcards)


# --- تابع نمایش متن فارسی ---
def persian_text(text):
    reshaped = arabic_reshaper.reshape(str(text))
    return get_display(reshaped)


# --- رسم نمودار ---
plt.figure(figsize=(12, 6))
bars = plt.bar(
    [persian_text(p) for p in provinces],
    simcards,
    color='#2196F3'  # آبی
)

# --- تنظیمات نمودار ---
plt.title(persian_text('توزیع سیم‌کارت در استان‌ها'), pad=20)
plt.xticks(rotation=45, ha='right')

# --- نمایش اعداد روی ستون‌ها ---
for bar in bars:
    height = bar.get_height()
    plt.text(
        bar.get_x() + bar.get_width() / 2,
        height,
        f'{height:,}',
        ha='center',
        va='bottom'
    )

# --- نمایش مجموع در پایین ---
footer = persian_text(f'جمع کل: {total:,} سیم‌کارت')
plt.figtext(
    0.5, 0.01,
    footer,
    ha='center',
    fontsize=12,
    bbox={'facecolor': 'white', 'pad': 5}
)

plt.tight_layout()
plt.show()